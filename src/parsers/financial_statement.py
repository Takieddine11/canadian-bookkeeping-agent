"""Parser for QBO **Balance Sheet** and **Profit and Loss** Excel exports.

Both reports share the same vertical-tree shape::

    Row 1: <Company>
    Row 2: <Report type>                 (e.g., "Balance Sheet", "Profit and Loss")
    Row 3: <Period>                      (e.g., "As of December 31, 2025",
                                                "January - December 2025")
    Row 4: <blank>
    Row 5:         | Total | <period columns if multi-period>
    ...
    Row N: <Section Header>              (no value, all-caps or Title-cased)
    Row N+1: ...  <Account>              (3-space indent per level; leaf = value)
    Row N+2: ...  Total <Section>        (rollup line, has a value)
    ...
    Row M: Tuesday, Apr. 21, 2026 ... - Accrual Basis

Indentation in column A is the only hierarchy signal QBO exposes — 3 spaces per level.
"""

from __future__ import annotations

import logging
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

log = logging.getLogger(__name__)

INDENT_PER_LEVEL = 3
# Matches the footer-basis string. Must anchor on the full phrase — "Cash"
# alone is a legitimate account name and must not be treated as a basis marker.
_BASIS_RE = re.compile(
    r"\b(Accrual|Cash)\s*Basis\b"
    r"|\b(Comptabilit[ée]\s*d'exercice)\b"
    r"|\b(Base\s*de\s*caisse)\b"
    r"|\b(Comptabilit[ée]\s*de\s*caisse)\b",
    re.IGNORECASE,
)

_FR_ACCRUAL_KEYWORDS = ("exercice",)


def _basis_label(match: re.Match) -> str:
    """Extract a human-readable basis label from a _BASIS_RE match."""
    text = (match.group(0) or "").lower()
    if "accrual" in text or any(k in text for k in _FR_ACCRUAL_KEYWORDS):
        return "Accrual"
    if "cash" in text or "caisse" in text:
        return "Cash"
    return match.group(0).capitalize()
# "As of …" (EN) | "Au …" (FR short) | "En date du …" (FR long) | "Du … au …" (FR range)
_AS_OF_RE = re.compile(
    r"(?:As\s*of|En\s*date\s*du|Au)\s+(.+)",
    re.IGNORECASE,
)
_PERIOD_DATE_FORMATS = (
    "%B %d, %Y",     # "December 31, 2025"
    "%b %d, %Y",     # "Dec 31, 2025"
    "%d %B %Y",      # "31 décembre 2025"  (FR long month, no comma)
    "%d %B, %Y",     # "31 décembre, 2025"
    "%d %b %Y",      # "31 déc 2025"
    "%d %b, %Y",     # "31 déc, 2025"
    "%d %b. %Y",     # "31 déc. 2025"
    "%d %b., %Y",    # "31 déc., 2025"
    "%Y-%m-%d",
)

# French month names: strptime on Windows respects the system locale, which
# is unreliable. Pre-normalize FR month names to their EN equivalents so we
# can parse with the default C locale.
_FR_MONTH_MAP = {
    "janvier": "January", "janv.": "Jan", "janv": "Jan",
    "février": "February", "fevrier": "February", "févr.": "Feb", "fevr.": "Feb",
    "févr": "Feb", "fevr": "Feb",
    "mars": "March",
    "avril": "April", "avr.": "Apr", "avr": "Apr",
    "mai": "May",
    "juin": "June",
    "juillet": "July", "juil.": "Jul", "juil": "Jul",
    "août": "August", "aout": "August",
    "septembre": "September", "sept.": "Sep", "sept": "Sep",
    "octobre": "October", "oct.": "Oct", "oct": "Oct",
    "novembre": "November", "nov.": "Nov", "nov": "Nov",
    "décembre": "December", "decembre": "December",
    "déc.": "Dec", "dec.": "Dec", "déc": "Dec", "dec": "Dec",
}


def _normalize_fr_months(text: str) -> str:
    """Replace French month names (including common abbreviations) with their
    English equivalents so that ``datetime.strptime`` can parse them without
    relying on the system locale."""
    lowered = text.lower()
    for fr, en in _FR_MONTH_MAP.items():
        if fr in lowered:
            # case-insensitive replace, preserving rest of string
            pattern = re.compile(re.escape(fr), re.IGNORECASE)
            text = pattern.sub(en, text)
            lowered = text.lower()
    return text
REPORT_BALANCE_SHEET = "balance_sheet"
REPORT_PNL = "pnl"


@dataclass(frozen=True)
class StatementLine:
    name: str
    level: int
    amount: Decimal | None
    is_total: bool
    is_section: bool  # header row with no amount (e.g., "Assets", "EXPENSES")


@dataclass(frozen=True)
class FinancialStatement:
    company: str
    report_type: str        # "balance_sheet" | "pnl"
    report_title: str       # raw title from row 2
    period_label: str
    as_of: date | None
    basis: str              # "Accrual" | "Cash" | ""
    lines: list[StatementLine]

    def find(self, name: str) -> StatementLine | None:
        """First line whose trimmed name matches (case-insensitive).

        Some QBO French PDF exports render accounts as bilingual labels with
        a slash separator — e.g. ``"Retained Earnings / Bénéfices non
        répartis"``. To let callers look up by either language without
        hard-coding every slash-pair, we also accept matches on any segment
        split by ``" / "``.
        """
        target = name.strip().lower()
        for line in self.lines:
            full = line.name.lower()
            if full == target:
                return line
            if " / " in full:
                segments = [s.strip() for s in line.name.split(" / ")]
                if any(s.lower() == target for s in segments):
                    return line
        return None

    def amount_of(self, name: str) -> Decimal | None:
        """Convenience: amount of the first matching line, or None."""
        line = self.find(name)
        return line.amount if line else None

    def amount_of_any(self, *names: str) -> Decimal | None:
        """Try each name in order; return the first matching line's amount.

        Used to accept English-or-French labels without branching in every agent.
        E.g. ``bs.amount_of_any("Total Assets", "Total de l'actif")``.
        """
        for n in names:
            v = self.amount_of(n)
            if v is not None:
                return v
        return None


class StatementParseError(Exception):
    pass


def parse_balance_sheet(path: Path | str) -> FinancialStatement:
    fs = parse_financial_statement(path)
    if fs.report_type != REPORT_BALANCE_SHEET:
        raise StatementParseError(
            f"Not a Balance Sheet (got report_title={fs.report_title!r})"
        )
    return fs


def parse_pnl(path: Path | str) -> FinancialStatement:
    fs = parse_financial_statement(path)
    if fs.report_type != REPORT_PNL:
        raise StatementParseError(f"Not a P&L (got report_title={fs.report_title!r})")
    return fs


def parse_financial_statement(path: Path | str) -> FinancialStatement:
    """Auto-detect Balance Sheet vs P&L from the file's own title row.

    Routes to the xlsx or PDF backend based on file extension.
    """
    p = Path(path)
    ext = p.suffix.lower()
    if ext in {".xlsx", ".xlsm"}:
        return _parse_statement_xlsx(p)
    if ext == ".pdf":
        return _parse_statement_pdf(p)
    raise StatementParseError(f"Unsupported file extension for financial statement: {ext}")


# ---- xlsx backend -------------------------------------------------------------


def _parse_statement_xlsx(path: Path) -> FinancialStatement:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        raise StatementParseError(f"empty workbook: {path}")

    company = _first_str(rows, 0)
    title = _first_str(rows, 1)
    period_label = _first_str(rows, 2)
    report_type = _classify_report(title)

    as_of = _parse_as_of(period_label)
    basis = _find_basis(rows)

    data_start = _find_data_start(rows)
    lines = [ln for ln in (_build_line(r) for r in rows[data_start:]) if ln is not None]

    return FinancialStatement(
        company=company,
        report_type=report_type,
        report_title=title,
        period_label=period_label,
        as_of=as_of,
        basis=basis,
        lines=lines,
    )


# ---- PDF backend --------------------------------------------------------------

# QBO PDF exports use a right-aligned amount column at roughly x0 > 500 and a
# left account-name column starting at x0 ~= 28. Indentation steps are about
# 3pt per level. We snap to the nearest level using the narrowest observed
# ``x0`` as "level 0".
_PDF_AMOUNT_X_THRESHOLD = 400
_PDF_LEVEL_STEP = 3


def _parse_statement_pdf(path: Path) -> FinancialStatement:
    with pdfplumber.open(path) as pdf:
        all_lines: list[_PdfLine] = []
        for page in pdf.pages:
            words = page.extract_words()
            all_lines.extend(_group_words_into_lines(words))

    if not all_lines:
        raise StatementParseError(f"no text extracted from PDF: {path}")

    # Preamble detection. QBO PDF exports use different orderings depending
    # on the export path (English: Company, Title, Period; Quebec/FR: Title,
    # Company, Period), and the company name can also be rendered twice
    # (logo alt-text + text header). Rather than trust positional indices,
    # scan the first ~8 non-amount lines and use *content* to identify each
    # field: the title is a known report marker, the period has a 4-digit
    # year or "as of / au / en date du" prefix, and the company is what
    # remains.
    _PREAMBLE_SCAN = 8
    preamble: list[_PdfLine] = []
    for ln in all_lines:
        if ln.amount is None and len(preamble) < _PREAMBLE_SCAN:
            preamble.append(ln)
        else:
            break

    report_type: str | None = None
    title = ""
    title_idx: int | None = None
    for i, ln in enumerate(preamble):
        rt = _maybe_report_type(ln.name)
        if rt is not None:
            report_type = rt
            title = ln.name
            title_idx = i
            break

    if report_type is None:
        preview = " | ".join(ln.name for ln in preamble[:4]) or "(empty)"
        raise StatementParseError(
            f"Unknown report type — no title marker in preamble. Preview: {preview!r}"
        )

    # Split the remaining preamble into period-like vs company-like lines.
    period_label = ""
    company_parts: list[str] = []
    for i, ln in enumerate(preamble):
        if i == title_idx:
            continue
        if not period_label and _looks_like_period(ln.name):
            period_label = ln.name
            continue
        # Stop consuming once we've hit the period — anything after is body.
        if period_label:
            break
        if ln.name.strip():
            company_parts.append(ln.name)
    company = " ".join(company_parts).strip()

    as_of = _parse_as_of(period_label)
    # Header = every preamble row we consumed as title/company/period.
    # Anything after that in `all_lines` is body — even if still non-amount
    # (section headers like "Actifs").
    consumed_idx = title_idx
    for i, ln in enumerate(preamble):
        if i == title_idx:
            continue
        if ln.name == period_label and period_label:
            consumed_idx = max(consumed_idx, i)
        elif ln.name in company_parts:
            consumed_idx = max(consumed_idx, i)
    header_lines = preamble[: consumed_idx + 1]

    # Footer with basis (and the "TOTAL" column header) are non-data lines to skip.
    body = all_lines[len(header_lines):]
    basis = ""
    cleaned: list[_PdfLine] = []
    for ln in body:
        if _BASIS_RE.search(ln.name):
            m = _BASIS_RE.search(ln.name)
            if m and not basis:
                basis = _basis_label(m)
            continue
        if ln.name.strip().upper() == "TOTAL" and ln.amount is None:
            continue  # column header
        if _is_page_footer_line(ln):
            continue
        cleaned.append(ln)

    if not cleaned:
        raise StatementParseError(f"no data lines in PDF: {path}")

    min_x = min(ln.x0 for ln in cleaned)
    lines: list[StatementLine] = []
    for ln in cleaned:
        level = max(0, round((ln.x0 - min_x) / _PDF_LEVEL_STEP))
        name = ln.name.strip()
        if not name:
            continue
        is_total = name.lower().startswith("total ") or name.lower() == "total"
        is_section = ln.amount is None and not is_total
        lines.append(StatementLine(
            name=name, level=level, amount=ln.amount,
            is_total=is_total, is_section=is_section,
        ))

    return FinancialStatement(
        company=company,
        report_type=report_type,
        report_title=title,
        period_label=period_label,
        as_of=as_of,
        basis=basis,
        lines=lines,
    )


# Common page-footer garbage QBO bakes into the PDF: page numbers ("1/1",
# "Page 1 of 2") and weekday timestamps ("mercredi, 22 avril 2026 11:51 PM
# GMTZ"). These have no amount and sit far to the right, so they pollute
# the body with bogus deeply-indented "section headers" if not filtered.
_PAGE_NUM_RE = re.compile(r"^\s*(Page\s+)?\d+\s*/\s*\d+\s*$", re.IGNORECASE)
_WEEKDAY_RE = re.compile(
    r"\b(monday|tuesday|wednesday|thursday|friday|saturday|sunday"
    r"|lundi|mardi|mercredi|jeudi|vendredi|samedi|dimanche)\b",
    re.IGNORECASE,
)


def _is_page_footer_line(ln: "_PdfLine") -> bool:
    if ln.amount is not None:
        return False
    n = ln.name.strip()
    if not n:
        return True
    if _PAGE_NUM_RE.match(n):
        return True
    if _WEEKDAY_RE.search(n):
        return True
    return False


@dataclass(frozen=True)
class _PdfLine:
    """One reconstructed row from a PDF: name (left) + optional amount (right)."""

    name: str
    amount: Decimal | None
    x0: float  # left-most x of the name; used to infer indent level


def _group_words_into_lines(words: list[dict]) -> list[_PdfLine]:
    """Cluster pdfplumber words by y-coordinate into lines, splitting name/amount by x.

    The amount can arrive as one token ("$12,345.67"), two tokens ("12,345.67 $"
    for French QBO), or even three tokens when the PDF uses a thin space as a
    thousands separator ("86 946,39 $"). We simply take every word whose left
    edge sits past the amount-column threshold, join them, and coerce — that
    covers all three layouts without special-casing.
    """
    if not words:
        return []

    buckets: dict[int, list[dict]] = defaultdict(list)
    for w in words:
        buckets[round(w["top"])].append(w)

    out: list[_PdfLine] = []
    for y in sorted(buckets):
        sorted_words = sorted(buckets[y], key=lambda w: w["x0"])
        name_parts: list[str] = []
        amount_parts: list[str] = []
        for w in sorted_words:
            if w["x0"] >= _PDF_AMOUNT_X_THRESHOLD:
                amount_parts.append(w["text"])
            else:
                name_parts.append(w["text"])

        amount: Decimal | None = None
        if amount_parts:
            joined = "".join(amount_parts)  # "12 345,67$" → "12345,67$"
            if _looks_like_amount(joined):
                amount = _coerce_amount(joined)

        # If the right-column tokens didn't form a valid amount, they're not
        # really an amount (e.g., a TOTAL column header, or stray text);
        # keep them in the name.
        if amount is None and amount_parts:
            name_parts.extend(amount_parts)

        if not name_parts and amount is None:
            continue
        name = " ".join(name_parts)
        x0 = sorted_words[0]["x0"]
        out.append(_PdfLine(name=name, amount=amount, x0=x0))
    return out


# Matches EN and FR numeric formats after we concatenate right-column tokens:
#   $1,234.56        EN (comma thousands, dot decimal)
#   (1,234.56)       EN negative in parens
#   12,345.67        EN no currency
#   -12,63$          FR (comma decimal, trailing $)
#   86946,39$        FR
#   1234567,89       FR no currency
# "\d{2,3}" decimal covers both "2 decimals" (currency) and "a few" to be
# tolerant of odd exports.
_AMOUNT_LIKE_RE = re.compile(
    r"^[-\$]?-?[0-9.,\s]+[.,]\d{2}\$?$"
    r"|^\([0-9.,\s]+[.,]\d{2}\)\$?$"
)


def _looks_like_amount(s: str) -> bool:
    return bool(_AMOUNT_LIKE_RE.match(s.strip()))


def _first_str(rows: list[tuple], idx: int) -> str:
    if idx >= len(rows):
        return ""
    cell = rows[idx][0] if rows[idx] else None
    return str(cell).strip() if cell is not None else ""


_PERIOD_MARKERS = (
    "as of",
    "for the period",
    "au ",                # "Au 31 décembre 2025"
    "en date du",         # "En date du 31 déc., 2025"
    "du ",                # "Du 1er janvier au 31 décembre"
    "pour la période", "pour la periode",
    "january", "february", "march", "april", "may", "june", "july",
    "august", "september", "october", "november", "december",
    "janvier", "février", "fevrier", "mars", "avril", "mai", "juin",
    "juillet", "août", "aout", "septembre", "octobre", "novembre", "décembre", "decembre",
)

_YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")


def _looks_like_period(text: str) -> bool:
    t = (text or "").lower()
    if _YEAR_RE.search(t):
        return True
    return any(m in t for m in _PERIOD_MARKERS)


_BS_TITLE_MARKERS = (
    "balance sheet",
    "bilan",
    "situation financière", "situation financiere",
    "état de situation", "etat de situation",
)
_PNL_TITLE_MARKERS = (
    "profit and loss", "income statement",
    "résultat", "resultat",
    "état des résultats", "etat des resultats",
    "état du résultat", "etat du resultat",
    "profits et pertes", "pertes et profits",
)


def _maybe_report_type(text: str) -> str | None:
    t = (text or "").lower()
    if any(m in t for m in _BS_TITLE_MARKERS):
        return REPORT_BALANCE_SHEET
    if any(m in t for m in _PNL_TITLE_MARKERS):
        return REPORT_PNL
    return None


def _classify_report(title: str) -> str:
    """Detect Balance Sheet vs P&L from the title row. Accepts English and French
    QBO labels — Quebec bookkeepers commonly export in French."""
    result = _maybe_report_type(title)
    if result is None:
        raise StatementParseError(f"Unknown report type in title: {title!r}")
    return result


def _parse_as_of(period_label: str) -> date | None:
    label = _normalize_fr_months(period_label)
    m = _AS_OF_RE.match(label)
    if m:
        raw = m.group(1).strip().rstrip(".,")
        for fmt in _PERIOD_DATE_FORMATS:
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
    # P&L period like "January - December 2025" or "January-December 2025"
    # (FR often drops the spaces): return Dec 31 of that year.
    year_match = re.search(r"\b(\d{4})\b", label)
    if year_match and re.search(r"\b(January|February|March|April|May|June|July|"
                                r"August|September|October|November|December)\b",
                                label, re.IGNORECASE):
        return date(int(year_match.group(1)), 12, 31)
    return None


def _find_basis(rows: list[tuple]) -> str:
    for row in rows[-15:]:  # basis is always in the footer
        for cell in row:
            if isinstance(cell, str):
                m = _BASIS_RE.search(cell)
                if m:
                    return _basis_label(m)
    return ""


def _find_data_start(rows: list[tuple]) -> int:
    """Return the row index after the column-header row (the one with 'Total')."""
    for i, row in enumerate(rows[:10]):
        if row and any(isinstance(c, str) and c.strip() == "Total" for c in row[1:]):
            return i + 1
    # fallback: skip the 3 preamble rows + 1 blank
    return 4


def _build_line(row: tuple) -> StatementLine | None:
    if not row or row[0] is None:
        return None
    name_raw = row[0]
    if not isinstance(name_raw, str):
        return None

    # Footer row (timestamp + basis) is text with a weekday/month. Skip it.
    if _BASIS_RE.search(name_raw):
        return None

    indent = len(name_raw) - len(name_raw.lstrip(" "))
    level = indent // INDENT_PER_LEVEL
    name = name_raw.strip()
    if not name:
        return None

    amount = _coerce_amount(row[1] if len(row) > 1 else None)
    is_total = name.lower().startswith("total ") or name.lower() == "total"
    is_section = amount is None and not is_total

    return StatementLine(
        name=name, level=level, amount=amount, is_total=is_total, is_section=is_section
    )


def _coerce_amount(cell) -> Decimal | None:  # noqa: ANN001 — openpyxl cell value
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return Decimal(str(cell))
    if isinstance(cell, Decimal):
        return cell
    if isinstance(cell, str):
        s = cell.strip()
        if not s:
            return None
        negative = s.startswith("(") and s.endswith(")")
        if negative:
            s = s[1:-1]
        # Strip currency + whitespace first.
        s = s.replace("$", "").replace(" ", "").replace("\xa0", "")
        # Handle EN vs FR number formats. EN: "1,234.56" (comma thousands,
        # dot decimal). FR: "1 234,56" or "1.234,56" (space/dot thousands,
        # comma decimal). After stripping spaces we can infer by which
        # separator appears last: that one is the decimal.
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                # FR: "1.234,56" — dot is thousands, comma is decimal.
                s = s.replace(".", "").replace(",", ".")
            else:
                # EN: "1,234.56" — comma is thousands, dot is decimal.
                s = s.replace(",", "")
        elif "," in s:
            # Only comma present. Treat as decimal if followed by 1-3 digits
            # at the end of the string (FR convention) — otherwise it's a
            # thousands separator (EN "1,234" = 1234).
            tail = s.rsplit(",", 1)[-1]
            if 1 <= len(tail) <= 3 and tail.isdigit() and len(tail) != 3:
                # FR decimal: "12,63" or "86946,3"
                s = s.replace(",", ".")
            elif 1 <= len(tail) <= 3 and tail.isdigit() and len(tail) == 3:
                # Ambiguous: "1,234" could be 1.234 (FR) or 1234 (EN thousands).
                # QBO never emits bare 3-digit "thousands" without a decimal,
                # so treat as FR decimal.
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")
        # If only "." present, leave as-is (EN decimal).
        try:
            value = Decimal(s)
            return -value if negative else value
        except Exception:
            return None
    return None
